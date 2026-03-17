#!/usr/bin/env python3
"""
Creates the sample Excel data files for the bootcamp scheduling practical.

Run once during setup:
    python3 create_data.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import os

OUT_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(OUT_DIR, exist_ok=True)

# ── Dates ──────────────────────────────────────────────────────────────────
# 4 weeks, Mon 16 Mar 2026 – Fri 10 Apr 2026

start = date(2026, 3, 16)
dates = [start + timedelta(days=w * 7 + d) for w in range(4) for d in range(5)]


# ── Availability ───────────────────────────────────────────────────────────
# (week 0-3, weekday 0=Mon..4=Fri)

def is_available(trainer, d):
    w = (d - start).days // 7  # 0–3
    wd = d.weekday()            # 0=Mon..4=Fri

    if trainer == "Alice Chen":
        return (w, wd) in {(0,0),(0,1),(0,3),(0,4),
                           (1,3),(1,4),
                           (2,0),(2,1),(2,3),
                           (3,3),(3,4)}
    if trainer == "Bob Smith":
        return (w, wd) in {(0,0),(0,1),
                           (1,0),(1,1),(1,3),(1,4),
                           (2,0),(2,1),
                           (3,0),(3,1)}
    if trainer == "Carol Jones":
        return (w, wd) in {(0,3),(0,4),
                           (1,0),(1,1),
                           (2,3),
                           (3,0),(3,1),(3,3),(3,4)}
    if trainer == "Diana Müller":
        return (w, wd) in {(0,0),(0,1),
                           (1,0),(1,1),(1,3),(1,4),
                           (2,3),(2,4),
                           (3,1),(3,3),(3,4)}
    if trainer == "Eve Brown":
        return (w, wd) in {(0,3),(0,4),
                           (1,0),(1,1),
                           (2,3),
                           (3,0),(3,1)}
    return False


TRAINERS = ["Alice Chen", "Bob Smith", "Carol Jones", "Diana Müller", "Eve Brown"]

# Trainer metadata (for intermediate / advanced)
TRAINER_META = {
    "Alice Chen":   {"home": "London",      "french": "No",  "max_week": None},
    "Bob Smith":    {"home": "Manchester",  "french": "No",  "max_week": 2},
    "Carol Jones":  {"home": "London",      "french": "No",  "max_week": None},
    "Diana Müller": {"home": "Paris",       "french": "Yes", "max_week": None},
    "Eve Brown":    {"home": "Bristol",     "french": "No",  "max_week": None},
}

# Weekly caps (ISO week number -> max bootcamps that week)
WEEKLY_CAPS = {
    (2026, 12): 2,   # week of 16 Mar
    (2026, 13): 2,   # week of 23 Mar
    (2026, 14): 1,   # week of 30 Mar (Easter week — reduced capacity)
    (2026, 15): 2,   # week of 6 Apr
}

# ── Style helpers ──────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CAP_FILL = PatternFill("solid", fgColor="FFF2CC")
YES_FILL = PatternFill("solid", fgColor="E2EFDA")
BOLD = Font(bold=True)


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")


def style_yes(cell):
    cell.fill = YES_FILL
    cell.alignment = Alignment(horizontal="center")


# ── basic.xlsx ─────────────────────────────────────────────────────────────

def create_basic():
    wb = Workbook()
    ws = wb.active
    ws.title = "Availability"

    # Row 1: header — "Name" then date strings
    ws.cell(1, 1, "Name")
    style_header(ws.cell(1, 1))
    for col, d in enumerate(dates, start=2):
        cell = ws.cell(1, col, d.strftime("%Y-%m-%d"))
        style_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = 12

    ws.column_dimensions["A"].width = 16

    # Rows 2+: trainer availability
    for row, trainer in enumerate(TRAINERS, start=2):
        ws.cell(row, 1, trainer).font = BOLD
        for col, d in enumerate(dates, start=2):
            if is_available(trainer, d):
                cell = ws.cell(row, col, "Yes")
                style_yes(cell)

    path = os.path.join(OUT_DIR, "basic.xlsx")
    wb.save(path)
    print(f"Created {path}")


# ── intermediate.xlsx ──────────────────────────────────────────────────────

def create_intermediate():
    wb = Workbook()
    ws = wb.active
    ws.title = "Availability"

    DATE_START_COL = 5  # cols: Name, Home Location, French, Max/Week, dates...

    # Row 1: column headers
    for col, label in enumerate(["Name", "Home Location", "French", "Max/Week"], start=1):
        cell = ws.cell(1, col, label)
        style_header(cell)
    for col, d in enumerate(dates, start=DATE_START_COL):
        cell = ws.cell(1, col, d.strftime("%Y-%m-%d"))
        style_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = 12

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10

    # Row 2: cap row (blank name, weekly caps on Monday columns)
    cap_cell = ws.cell(2, 1, "")  # blank name
    cap_cell.fill = CAP_FILL
    for col, d in enumerate(dates, start=DATE_START_COL):
        cell = ws.cell(2, col)
        cell.fill = CAP_FILL
        if d.weekday() == 0:  # Monday
            yr, wk, _ = d.isocalendar()
            cap = WEEKLY_CAPS.get((yr, wk))
            if cap:
                cell.value = cap
                cell.alignment = Alignment(horizontal="center")

    # Rows 3+: trainer rows
    for row, trainer in enumerate(TRAINERS, start=3):
        meta = TRAINER_META[trainer]
        ws.cell(row, 1, trainer).font = BOLD
        ws.cell(row, 2, meta["home"])
        ws.cell(row, 3, meta["french"])
        if meta["max_week"]:
            ws.cell(row, 4, meta["max_week"])
        for col, d in enumerate(dates, start=DATE_START_COL):
            if is_available(trainer, d):
                cell = ws.cell(row, col, "Yes")
                style_yes(cell)

    path = os.path.join(OUT_DIR, "intermediate.xlsx")
    wb.save(path)
    print(f"Created {path}")


# ── advanced.xlsx ──────────────────────────────────────────────────────────

LOCATIONS = [
    {"name": "London",    "country": "UK",          "demand": 3, "french": "No",  "max_parallel": 1, "max_week": None},
    {"name": "Paris",     "country": "France",      "demand": 2, "french": "Yes", "max_parallel": 1, "max_week": None},
    {"name": "Amsterdam", "country": "Netherlands", "demand": 1, "french": "No",  "max_parallel": 1, "max_week": None},
    {"name": "Stockholm", "country": "Sweden",      "demand": 1, "french": "No",  "max_parallel": 1, "max_week": None},
]

WEIGHTINGS = {
    "Alice Chen":   3,
    "Bob Smith":    2,
    "Carol Jones":  1,
    "Diana Müller": 3,
    "Eve Brown":    1,
}


def create_advanced():
    wb = Workbook()

    # ── Availability sheet (same as intermediate) ──
    ws = wb.active
    ws.title = "Availability"
    DATE_START_COL = 5

    for col, label in enumerate(["Name", "Home Location", "French", "Max/Week"], start=1):
        cell = ws.cell(1, col, label)
        style_header(cell)
    for col, d in enumerate(dates, start=DATE_START_COL):
        cell = ws.cell(1, col, d.strftime("%Y-%m-%d"))
        style_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = 12

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10

    cap_cell = ws.cell(2, 1, "")
    cap_cell.fill = CAP_FILL
    for col, d in enumerate(dates, start=DATE_START_COL):
        cell = ws.cell(2, col)
        cell.fill = CAP_FILL
        if d.weekday() == 0:
            yr, wk, _ = d.isocalendar()
            cap = WEEKLY_CAPS.get((yr, wk))
            if cap:
                cell.value = cap
                cell.alignment = Alignment(horizontal="center")

    for row, trainer in enumerate(TRAINERS, start=3):
        meta = TRAINER_META[trainer]
        ws.cell(row, 1, trainer).font = BOLD
        ws.cell(row, 2, meta["home"])
        ws.cell(row, 3, meta["french"])
        if meta["max_week"]:
            ws.cell(row, 4, meta["max_week"])
        for col, d in enumerate(dates, start=DATE_START_COL):
            if is_available(trainer, d):
                cell = ws.cell(row, col, "Yes")
                style_yes(cell)

    # ── Locations sheet ──
    ws_loc = wb.create_sheet("Locations")
    headers = ["Name", "Country", "Demand", "French Required", "Max Parallel", "Max/Week"]
    for col, h in enumerate(headers, start=1):
        cell = ws_loc.cell(1, col, h)
        style_header(cell)
        ws_loc.column_dimensions[get_column_letter(col)].width = 16

    for row, loc in enumerate(LOCATIONS, start=2):
        ws_loc.cell(row, 1, loc["name"])
        ws_loc.cell(row, 2, loc["country"])
        ws_loc.cell(row, 3, loc["demand"])
        ws_loc.cell(row, 4, loc["french"])
        ws_loc.cell(row, 5, loc["max_parallel"])
        if loc["max_week"]:
            ws_loc.cell(row, 6, loc["max_week"])

    # ── Weightings sheet ──
    ws_wt = wb.create_sheet("Weightings")
    ws_wt.cell(1, 1, "Name").font = BOLD
    ws_wt.cell(1, 2, "Weight").font = BOLD
    ws_wt.column_dimensions["A"].width = 16
    ws_wt.column_dimensions["B"].width = 10

    for row, (name, weight) in enumerate(WEIGHTINGS.items(), start=2):
        ws_wt.cell(row, 1, name)
        ws_wt.cell(row, 2, weight)

    path = os.path.join(OUT_DIR, "advanced.xlsx")
    wb.save(path)
    print(f"Created {path}")


# ── Run all ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    create_basic()
    create_intermediate()
    create_advanced()
    print("Done. Data files written to data/")
