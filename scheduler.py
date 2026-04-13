# scheduler.py
#
# Build your bootcamp scheduler here.
#
# Work through the challenges in order:
#   Run /next in Claude Code to see what to do next.
#   Run /spec to write a specification for the next function.
#   Run /build to implement from your spec.
#   Run python3 -m pytest tests/ -v to verify what you've built.
#
# Each challenge adds new functions — don't remove old ones.

from datetime import date, timedelta
import holidays
import openpyxl
from ortools.sat.python import cp_model

CITY_TO_COUNTRY_CODE = {
    'London': 'GB', 'Manchester': 'GB', 'Bristol': 'GB',
    'Paris': 'FR', 'Amsterdam': 'NL', 'Stockholm': 'SE',
}


def parse_availability(filepath: str) -> tuple:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Availability']

    # Detect extra columns (Home Location, French, Max/Week)
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    has_extra = 'Home Location' in headers or 'French' in headers or 'Max/Week' in headers

    # Find date columns
    dates = []
    date_cols = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val is None:
            continue
        val_str = str(val).strip()
        try:
            d = date.fromisoformat(val_str)
            dates.append(d)
            date_cols.append(col)
        except ValueError:
            pass
    dates_sorted = sorted(dates)
    date_to_idx = {d: i for i, d in enumerate(dates)}

    # Find extra column indices
    loc_col = headers.index('Home Location') + 1 if 'Home Location' in headers else None
    french_col = headers.index('French') + 1 if 'French' in headers else None
    max_col = headers.index('Max/Week') + 1 if 'Max/Week' in headers else None

    # Parse trainer rows
    trainers: dict[str, dict[date, bool]] = {}
    trainer_info: dict[str, dict] = {}
    weekly_caps: dict[tuple[int, int], int] = {}

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if not name or not str(name).strip():
            # Cap row: blank name, numbers on Monday columns = weekly caps
            if has_extra:
                for i, col in enumerate(date_cols):
                    val = ws.cell(row, col).value
                    if val is not None:
                        try:
                            cap = int(val)
                            d = dates[i]
                            iso_year, iso_week, _ = d.isocalendar()
                            weekly_caps[(iso_year, iso_week)] = cap
                        except (ValueError, TypeError):
                            pass
            continue

        name = str(name).strip()
        avail: dict[date, bool] = {}
        for i, col in enumerate(date_cols):
            cell_val = ws.cell(row, col).value
            avail[dates[i]] = (str(cell_val).strip().lower() == 'yes') if cell_val else False
        trainers[name] = avail

        # Parse extra columns
        if has_extra:
            info: dict = {}
            if loc_col:
                info['home_location'] = str(ws.cell(row, loc_col).value or '').strip()
            if french_col:
                fval = ws.cell(row, french_col).value
                info['french_speaking'] = str(fval).strip().lower() == 'yes' if fval else False
            if max_col:
                mval = ws.cell(row, max_col).value
                info['max_per_week'] = int(mval) if mval is not None else None
            trainer_info[name] = info

    if has_extra:
        return (dates_sorted, trainers, trainer_info, weekly_caps)
    return (dates_sorted, trainers)


def parse_locations(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Locations']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    locations = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if not name:
            continue
        loc = {
            'name': str(name).strip(),
            'country': str(ws.cell(row, headers.index('Country') + 1).value or '').strip(),
            'demand': int(ws.cell(row, headers.index('Demand') + 1).value or 0),
            'french_required': str(ws.cell(row, headers.index('French Required') + 1).value or '').strip().lower() == 'yes',
        }
        locations.append(loc)
    return locations


def apply_bank_holidays(trainers: dict[str, dict[date, bool]], trainer_info: dict[str, dict]) -> None:
    all_dates = set()
    for avail in trainers.values():
        all_dates.update(avail.keys())
    years = {d.year for d in all_dates}

    for name, avail in trainers.items():
        info = trainer_info.get(name, {})
        home = info.get('home_location', '')
        country_code = CITY_TO_COUNTRY_CODE.get(home)
        if not country_code:
            continue
        country_holidays = holidays.country_holidays(country_code, years=years)
        for d in list(avail.keys()):
            if d in country_holidays:
                avail[d] = False


def parse_weightings(filepath: str) -> dict[str, int]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Weightings']
    result = {}
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value
        weight = ws.cell(row, 2).value
        if name and weight is not None:
            result[str(name).strip()] = int(weight)
    return result


DAY_MAP = {'mon': 0, 'tue': 1, 'wed': 2, 'thu': 3, 'fri': 4, 'sat': 5, 'sun': 6}


def generate_slots(
    dates: list[date],
    trainers: dict[str, dict[date, bool]],
    pattern: str = 'mon-tue,thu-fri',
) -> list[tuple[date, date]]:
    # Parse pattern into weekday pairs
    pairs = []
    for part in pattern.split(','):
        parts = part.strip().lower().split('-')
        pairs.append((DAY_MAP[parts[0]], DAY_MAP[parts[1]]))

    # Build date lookup by (year, week, weekday)
    date_set = set(dates)

    slots = []
    for d in sorted(dates):
        wd = d.weekday()
        for wd1, wd2 in pairs:
            if wd == wd1:
                d2 = d + timedelta(days=wd2 - wd1)
                if d2 not in date_set:
                    continue
                # Check at least 2 trainers available on both days
                count = sum(
                    1 for avail in trainers.values()
                    if avail.get(d, False) and avail.get(d2, False)
                )
                if count >= 2:
                    slots.append((d, d2))
    return slots


def schedule_greedy(
    dates: list[date],
    trainers: dict[str, dict[date, bool]],
    slots: list[tuple[date, date]],
    config: dict | None = None,
) -> list[dict]:
    booked: dict[date, set[str]] = {}
    schedule = []

    for slot in slots:
        d1, d2 = slot
        # Find trainers available on both days and not booked
        available = [
            name for name in sorted(trainers.keys())
            if trainers[name].get(d1, False)
            and trainers[name].get(d2, False)
            and name not in booked.get(d1, set())
            and name not in booked.get(d2, set())
        ]

        if config:
            experienced = config.get('experienced', set())
            trainees = config.get('trainees', set())
            # Try to find a valid pair: at least one experienced
            pair = _pick_pair_with_experience(available, experienced, trainees)
        else:
            pair = available[:2] if len(available) >= 2 else None

        if pair:
            schedule.append({'slot': slot, 'trainers': list(pair)})
            booked.setdefault(d1, set()).update(pair)
            booked.setdefault(d2, set()).update(pair)

    return schedule


def _pick_pair_with_experience(
    available: list[str],
    experienced: set[str],
    trainees: set[str],
) -> list[str] | None:
    # Need at least 2 available and at least 1 experienced
    if len(available) < 2:
        return None
    exp_avail = [n for n in available if n in experienced]
    if not exp_avail:
        return None
    # Pick first experienced trainer
    first = exp_avail[0]
    # Pick second: any non-trainee, or experienced
    # Trainees can only pair with experienced
    for second in available:
        if second == first:
            continue
        if second in trainees:
            # trainee can pair with experienced (first is experienced)
            return [first, second]
        else:
            # non-trainee or experienced can pair with experienced
            return [first, second]
    return None


def schedule_optimal(
    dates: list[date],
    trainers: dict[str, dict[date, bool]],
    slots: list[tuple[date, date]],
    config: dict | None = None,
    trainer_info: dict | None = None,
    weekly_caps: dict | None = None,
    locations: list[dict] | None = None,
    weightings: dict[str, int] | None = None,
) -> list[dict]:
    model = cp_model.CpModel()
    trainer_names = sorted(trainers.keys())

    # Boolean variables
    x = {}  # x[t,si] = trainer t assigned to slot si
    for t in trainer_names:
        for si, s in enumerate(slots):
            d1, d2 = s
            if trainers[t].get(d1, False) and trainers[t].get(d2, False):
                x[t, si] = model.new_bool_var(f'x_{t}_{si}')
            else:
                x[t, si] = model.new_constant(0)

    y = {}  # y[si] = slot si is active
    for si in range(len(slots)):
        y[si] = model.new_bool_var(f'y_{si}')

    # Each active slot gets exactly 2 trainers
    for si in range(len(slots)):
        model.add(sum(x[t, si] for t in trainer_names) == 2 * y[si])

    # No trainer double-booked on any day
    day_to_slots: dict[date, list[int]] = {}
    for si, s in enumerate(slots):
        d1, d2 = s
        day_to_slots.setdefault(d1, []).append(si)
        day_to_slots.setdefault(d2, []).append(si)

    for t in trainer_names:
        for d, slot_indices in day_to_slots.items():
            model.add(sum(x[t, si] for si in slot_indices) <= 1)

    # Experience rules
    if config:
        experienced = config.get('experienced', set())
        trainees_set = config.get('trainees', set())
        for si in range(len(slots)):
            model.add(
                sum(x[t, si] for t in trainer_names if t in experienced) >= y[si]
            )
            for t in trainer_names:
                if t in trainees_set:
                    model.add(x[t, si] <= sum(x[e, si] for e in trainer_names if e in experienced))

    # Weekly caps on trainers
    week_to_slots: dict[tuple[int, int], list[int]] = {}
    for si, s in enumerate(slots):
        d1 = s[0]
        iso_year, iso_week, _ = d1.isocalendar()
        week_to_slots.setdefault((iso_year, iso_week), []).append(si)

    if trainer_info:
        for t in trainer_names:
            if t in trainer_info and trainer_info[t].get('max_per_week') is not None:
                max_days = trainer_info[t]['max_per_week']
                max_bootcamps = max_days // 2
                for wk, sis in week_to_slots.items():
                    model.add(sum(x[t, si] for si in sis) <= max_bootcamps)

    if weekly_caps:
        for wk, cap in weekly_caps.items():
            if wk in week_to_slots:
                model.add(sum(y[si] for si in week_to_slots[wk]) <= cap)

    # Location assignment
    z = {}
    if locations:
        for si in range(len(slots)):
            for li in range(len(locations)):
                z[si, li] = model.new_bool_var(f'z_{si}_{li}')

        for si in range(len(slots)):
            model.add(sum(z[si, li] for li in range(len(locations))) == y[si])

        french_trainers = set()
        if trainer_info:
            french_trainers = {t for t in trainer_names if trainer_info.get(t, {}).get('french_speaking', False)}

        for li, loc in enumerate(locations):
            if loc.get('french_required', False):
                for si in range(len(slots)):
                    model.add(
                        sum(x[t, si] for t in french_trainers) >= z[si, li]
                    )

        demand_vars = []
        for li, loc in enumerate(locations):
            demand = loc.get('demand', 0)
            count = model.new_int_var(0, len(slots), f'count_{li}')
            model.add(count == sum(z[si, li] for si in range(len(slots))))
            capped = model.new_int_var(0, demand, f'capped_{li}')
            model.add_min_equality(capped, [count, model.new_constant(demand)])
            demand_vars.append(capped)

        objective_terms = [v * 1000 for v in demand_vars]

        if weightings:
            for t in trainer_names:
                w = weightings.get(t, 0)
                if w:
                    for si in range(len(slots)):
                        objective_terms.append(x[t, si] * w)

        # Travel penalties
        CITY_TO_COUNTRY = {
            'London': 'GB', 'Manchester': 'GB', 'Bristol': 'GB',
            'Paris': 'FR', 'Amsterdam': 'NL', 'Stockholm': 'SE',
        }
        COUNTRY_NAME_TO_ISO = {
            'UK': 'GB', 'France': 'FR', 'Netherlands': 'NL', 'Sweden': 'SE',
            'India': 'IN', 'USA': 'US', 'Singapore': 'SG', 'Australia': 'AU',
            'Japan': 'JP', 'China': 'CN', 'Brazil': 'BR', 'South Africa': 'ZA',
            'Germany': 'DE', 'Spain': 'ES', 'Italy': 'IT',
        }
        LONG_HAUL = {'IN', 'US', 'SG', 'AU', 'JP', 'CN', 'BR', 'ZA'}

        if trainer_info:
            for si in range(len(slots)):
                for li, loc in enumerate(locations):
                    raw_country = loc.get('country', '')
                    loc_country = COUNTRY_NAME_TO_ISO.get(raw_country, raw_country)
                    for t in trainer_names:
                        home = (trainer_info.get(t, {}).get('home_location', '') or '')
                        home_country = CITY_TO_COUNTRY.get(home, '')
                        if home_country and loc_country and home_country != loc_country:
                            penalty = 50 if (home_country in LONG_HAUL or loc_country in LONG_HAUL) else 10
                            both = model.new_bool_var(f'both_{t}_{si}_{li}')
                            model.add_bool_and([x[t, si], z[si, li]]).only_enforce_if(both)
                            model.add_bool_or([x[t, si].negated(), z[si, li].negated()]).only_enforce_if(both.negated())
                            objective_terms.append(-penalty * both)

        model.maximize(sum(objective_terms))
    else:
        objective_terms = [y[si] * 1000 for si in range(len(slots))]
        if weightings:
            for t in trainer_names:
                w = weightings.get(t, 0)
                if w:
                    for si in range(len(slots)):
                        objective_terms.append(x[t, si] * w)
        model.maximize(sum(objective_terms))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 30
    status = solver.solve(model)

    schedule = []
    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        for si, s in enumerate(slots):
            if solver.value(y[si]):
                assigned = [t for t in trainer_names if solver.value(x[t, si])]
                entry = {'slot': s, 'trainers': assigned}
                if locations:
                    for li, loc in enumerate(locations):
                        if solver.value(z[si, li]):
                            entry['location'] = loc['name']
                            break
                schedule.append(entry)

    return schedule
